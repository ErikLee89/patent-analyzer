import sys
import json
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

import re

def main():
    # Expecting: python save_report.py <json_file_path> <output_xlsx_path>
    if len(sys.argv) < 3:
        print("Usage: python save_report.py <json_file_path> <output_xlsx_path>")
        sys.exit(1)

    json_path = sys.argv[1]
    output_path = sys.argv[2]

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        df = pd.DataFrame(data)

        # --- Deduplication Logic: Prefer Granted (B) over Application (A) ---
        if not df.empty and "专利号" in df.columns:
            # Helper to extract base ID (remove trailing letters)
            def get_base_id(p_no):
                if not isinstance(p_no, str): return str(p_no)
                clean = p_no.replace(' ', '').strip()
                # Remove trailing letters (A, B, U, A1, B2, etc.)
                return re.sub(r'[A-Z]+\d*$', '', clean)

            # Helper for priority (Lower is better)
            def get_priority(p_no):
                if not isinstance(p_no, str): return 99
                clean = p_no.replace(' ', '').strip().upper()
                if 'B' in clean: return 1 # Granted
                return 2 # Others (A, U, etc.)

            df['_base_id'] = df['专利号'].apply(get_base_id)
            df['_priority'] = df['专利号'].apply(get_priority)
            
            # Sort by Base ID then Priority
            df = df.sort_values(by=['_base_id', '_priority'])
            
            # Drop duplicates on Base ID, keeping the first (highest priority)
            initial_count = len(df)
            df = df.drop_duplicates(subset=['_base_id'], keep='first')
            final_count = len(df)
            
            if initial_count > final_count:
                print(f"Filtered out {initial_count - final_count} application versions in favor of granted versions.")
            
            # Clean up temp columns
            df = df.drop(columns=['_base_id', '_priority'])

        # Add "序号" column at the beginning
        df.insert(0, "序号", range(1, len(df) + 1))

        # Reorder columns: Summary before 摘要
        preferred_order = ["序号", "专利号", "标题", "申请人", "发明人", "申请日", "授权日", "IPC分类号", "Summary", "摘要", "主权项"]
        existing_cols = [c for c in preferred_order if c in df.columns]
        other_cols = [c for c in df.columns if c not in existing_cols]
        df = df[existing_cols + other_cols]
        
        # Ensure directory exists
        output_dir = os.path.dirname(os.path.abspath(output_path))
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        
        # Save initially with pandas
        df.to_excel(output_path, index=False)
        
        # Open with openpyxl for formatting
        wb = load_workbook(output_path)
        ws = wb.active
        
        # --- Formatting Configuration ---
        # Fixed widths (approximate character width)
        fixed_widths = {
            "标题": 15,
            "申请人": 15,
            "发明人": 15,
            "Summary": 20,
            "摘要": 80,
            "主权项": 80
        }
        
        # Columns to auto-fit (we will approximate this by checking content length)
        auto_fit_columns = ["序号", "专利号", "IPC分类号", "申请日", "授权日"]

        # Define border style
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        # Iterate through all columns
        for col_idx, column_cell in enumerate(ws[1], 1): # 1-based index
            column_name = column_cell.value
            col_letter = column_cell.column_letter
            
            # 1. Apply Fixed Widths
            if column_name in fixed_widths:
                ws.column_dimensions[col_letter].width = fixed_widths[column_name]
            
            # 2. Apply Auto-fit (Simple Approximation)
            elif column_name in auto_fit_columns:
                max_length = 0
                for cell in ws[col_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2 # Add padding
                ws.column_dimensions[col_letter].width = adjusted_width
            
            # Default width for others
            else:
                 ws.column_dimensions[col_letter].width = 15

        # 3. Apply Word Wrap, Auto Height, and Borders to ALL cells in the data range
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = thin_border

        # Save formatting changes
        wb.save(output_path)
        print(f"Successfully saved {len(data)} records to {output_path} with formatting.")
        
    except Exception as e:
        print(f"Error creating report: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
